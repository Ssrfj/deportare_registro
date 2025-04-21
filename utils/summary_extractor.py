import os
import glob
from openpyxl import load_workbook
import logging

def extract_paper1_summary(club_path):
    """
    クラブフォルダ内の01_登録基準確認用紙.xlsx から主要情報を抽出して返す
    """
    summary = {}
    pattern = os.path.join(club_path, "01_*確認用紙.xlsx")
    files = glob.glob(pattern)

    if not files:
        logging.warning(f"{club_path}: 登録基準確認用紙が見つかりませんでした")
        return summary

    file_path = files[0]
    try:
        wb = load_workbook(file_path)
        ws = wb.active

        summary = {
            "クラブ名（紙①）": ws["A1"].value,
            "代表者名（紙①）": ws["A2"].value,
            "申請区分（紙①）": ws["A3"].value,
            "TEL（紙①）": ws["A30"].value,
            "Email（紙①）": ws["A31"].value
        }

    except Exception as e:
        logging.exception(f"{club_path}: 登録基準確認用紙の内容抽出に失敗しました")

    return summary
