import os
import glob
import logging
from openpyxl import load_workbook
from config.paths import REQUIRED_FILES
from utils.log_writer import save_log
from utils.template_loader import (
    load_check_template_paper1,
    load_check_template_paper2_1,
    load_check_template_paper2_2,
    load_check_template_paper3
)

def check_club_info(ws, club_master_dict):
    club_name = ws["A1"].value
    application_type = ws["A3"].value
    messages = []

    if not club_name:
        messages.append("â— ã‚¯ãƒ©ãƒ–åãŒæœªè¨˜å…¥ã§ã™ã€‚")
    elif club_name not in club_master_dict:
        messages.append(f"â— ã‚¯ãƒ©ãƒ–åã€{club_name}ã€ã¯åç°¿ã«å­˜åœ¨ã—ã¾ã›ã‚“ã€‚")
    else:
        expected = club_master_dict[club_name]
        if application_type != expected:
            messages.append(f"âš ï¸ ç”³è«‹åŒºåˆ†ãŒä¸ä¸€è‡´ï¼šåç°¿ã§ã¯ã€{expected}ã€ã€ç”³è«‹æ›¸ã§ã¯ã€{application_type}ã€ã§ã™ã€‚")

    return messages

def check_cells_with_template(ws, template_rows, doc_name=""):
    prefix = f"[{doc_name}] " if doc_name else ""
    issues = []

    for item in template_rows:
        cell = item.get("ã‚»ãƒ«ä½ç½®")
        name = item.get("é …ç›®å")
        check_type = item.get("ãƒã‚§ãƒƒã‚¯ç¨®åˆ¥")
        required = item.get("å¿…é ˆã‹")
        notes = item.get("å‚™è€ƒ", "")

        # ç¯„å›²æŒ‡å®šã‚»ãƒ«ã®å‡¦ç†
        if ":" in cell:
            try:
                start_cell, end_cell = cell.split(":")
                for row in ws[start_cell:end_cell]:
                    for c in row:
                        val = c.value
                        if check_type == "â—‹ã®æœ‰ç„¡" and required == "â—‹":
                            valid_marks = {"â—‹", "æœ‰", "1", "âœ”", "yes", "Yes"}
                            if not val or str(val).strip() not in valid_marks:
                                issues.append(f"{prefix}â— {name}ï¼ˆ{c.coordinate}ï¼‰ã«â—‹å°ãŒã‚ã‚Šã¾ã›ã‚“ã€‚")
            except Exception as e:
                issues.append(f"{prefix}âŒ ã‚»ãƒ«ç¯„å›² {cell} ã®ãƒã‚§ãƒƒã‚¯ä¸­ã«ã‚¨ãƒ©ãƒ¼ãŒç™ºç”Ÿã—ã¾ã—ãŸã€‚")
            continue  # å€‹åˆ¥ã‚»ãƒ«å‡¦ç†ã‚’ã‚¹ã‚­ãƒƒãƒ—

        value = ws[cell].value if cell else None

        if check_type == "ç©ºæ¬„ãƒã‚§ãƒƒã‚¯" and required == "â—‹":
            if not value or str(value).strip() == "":
                issues.append(f"â— {prefix}â— {name}ï¼ˆ{cell}ï¼‰ãŒæœªè¨˜å…¥ã§ã™ã€‚")

        elif check_type == "â—‹ã®æœ‰ç„¡" and required == "â—‹":
            if not value or "â—‹" not in str(value):
                issues.append(f"â— {prefix}â— {name}ï¼ˆ{cell}ï¼‰ã«â—‹å°ãŒã‚ã‚Šã¾ã›ã‚“ã€‚")

        elif check_type == "å€¤ãƒã‚§ãƒƒã‚¯" and required == "â—‹":
            if str(value).strip() not in ["æ–°è¦", "æ›´æ–°"]:
                issues.append(f"â— {prefix}â— {name}ï¼ˆ{cell}ï¼‰ãŒã€Œæ–°è¦ï¼æ›´æ–°ã€ã«ãªã£ã¦ã„ã¾ã›ã‚“ã€‚")

        elif check_type == "å½¢å¼ãƒã‚§ãƒƒã‚¯":
            if "é›»è©±" in name and (not value or "-" not in str(value)):
                issues.append(f"{prefix}â— {name}ï¼ˆ{cell}ï¼‰ã®å½¢å¼ãŒä¸æ­£ã§ã™ã€‚")
            elif "ãƒ¡ãƒ¼ãƒ«" in name and (not value or "@" not in str(value)):
                issues.append(f"{prefix}â— {name}ï¼ˆ{cell}ï¼‰ãŒæ­£ã—ã„ãƒ¡ãƒ¼ãƒ«ã‚¢ãƒ‰ãƒ¬ã‚¹å½¢å¼ã§ã¯ã‚ã‚Šã¾ã›ã‚“ã€‚")
            elif "äººæ•°" in name:
                try:
                    num = int(value)
                    if num < 0:
                        issues.append(f"{prefix}â— {name}ï¼ˆ{cell}ï¼‰ãŒ0ä»¥ä¸Šã®æ•°å€¤ã§ã¯ã‚ã‚Šã¾ã›ã‚“ã€‚")
                except:
                    issues.append(f"{prefix}â— {name}ï¼ˆ{cell}ï¼‰ãŒæ•°å€¤ã¨ã—ã¦èª­ã¿å–ã‚Œã¾ã›ã‚“ã€‚")

    return issues

def check_excel_contents_paper1(club_path, club_name, club_master_dict):
    issues = []

    pattern = os.path.join(club_path, "01_*ç¢ºèªç”¨ç´™.xlsx")
    files = glob.glob(pattern)

    if not files:
        issues.append("â— ç™»éŒ²åŸºæº–ç¢ºèªç”¨ç´™ãŒè¦‹ã¤ã‹ã‚Šã¾ã›ã‚“ã€‚")
        return issues

    file_path = files[0]
    try:
        wb = load_workbook(file_path)
        ws = wb.active

        # åç°¿ã¨ã®ç…§åˆãƒã‚§ãƒƒã‚¯ï¼ˆA1, A3ï¼‰
        issues += check_club_info(ws, club_master_dict)

        # ãƒ†ãƒ³ãƒ—ãƒ¬ãƒ¼ãƒˆèª­ã¿è¾¼ã¿
        template = load_check_template_paper1()
        if template:
            issues += check_cells_with_template(ws, template, doc_name="ç™»éŒ²åŸºæº–ç¢ºèªç”¨ç´™")

    except Exception as e:
        logging.exception(f"{club_name}: Excelå†…å®¹ã®èª­ã¿å–ã‚Šä¸­ã«ã‚¨ãƒ©ãƒ¼")
        issues.append("âŒ Excelãƒ•ã‚¡ã‚¤ãƒ«ã®èª­ã¿å–ã‚Šã«å¤±æ•—ã—ã¾ã—ãŸã€‚")

    return issues

def check_excel_contents_paper2_1(club_path, club_name):
    issues = []

    # æ›¸é¡ãƒ•ã‚¡ã‚¤ãƒ«ã®æ¢ç´¢ï¼ˆ02-1ã§å§‹ã¾ã‚‹ãƒ•ã‚¡ã‚¤ãƒ«ï¼‰
    pattern = os.path.join(club_path, "02-1_*åŸºç¤æƒ…å ±æ›¸é¡.xlsx")
    files = glob.glob(pattern)

    if not files:
        issues.append("â— æ›¸é¡â‘¡-1ã€åŸºç¤æƒ…å ±æ›¸é¡ã€ãŒè¦‹ã¤ã‹ã‚Šã¾ã›ã‚“ã€‚")
        return issues

    file_path = files[0]
    try:
        wb = load_workbook(file_path)
        ws = wb.active

        # ãƒ†ãƒ³ãƒ—ãƒ¬ãƒ¼ãƒˆèª­ã¿è¾¼ã¿
        template = load_check_template_paper2_1("templates/check_paper2_1_template.xlsx")
        print(f"ğŸ” {club_name}: æ›¸é¡â‘¡-1 ãƒ†ãƒ³ãƒ—ãƒ¬ãƒ¼ãƒˆä»¶æ•° = {len(template)}")

        if template:
            issues += check_cells_with_template(ws, template, doc_name="åŸºç¤æƒ…å ±æ›¸é¡")

    except Exception as e:
        logging.exception(f"{club_name}: æ›¸é¡â‘¡-1ã®Excelèª­ã¿è¾¼ã¿ã‚¨ãƒ©ãƒ¼")
        issues.append("âŒ æ›¸é¡â‘¡-1ã®èª­ã¿å–ã‚Šã«å¤±æ•—ã—ã¾ã—ãŸã€‚")

    return issues

def check_excel_contents_paper2_2(club_path, club_name):
    issues = []

    # ãƒ•ã‚¡ã‚¤ãƒ«æ¢ç´¢
    pattern = os.path.join(club_path, "02-2_*.xlsx")
    files = glob.glob(pattern)

    if not files:
        issues.append("â— æ›¸é¡â‘¡-2ã€æ´»å‹•ãƒ»ãƒãƒã‚¸ãƒ£ãƒ¼æ›¸é¡ã€ãŒè¦‹ã¤ã‹ã‚Šã¾ã›ã‚“ã€‚")
        return issues

    file_path = files[0]
    try:
        wb = load_workbook(file_path)
        ws = wb.active

        # ãƒ†ãƒ³ãƒ—ãƒ¬ãƒ¼ãƒˆèª­ã¿è¾¼ã¿
        template = load_check_template_paper2_2("templates/check_paper2_2_template.xlsx")
        print(f"ğŸ” {club_name}: æ›¸é¡â‘¡-2 ãƒ†ãƒ³ãƒ—ãƒ¬ãƒ¼ãƒˆä»¶æ•° = {len(template)}")

        if template:
            issues += check_cells_with_template(ws, template, doc_name="æ´»å‹•ãƒ»ãƒãƒã‚¸ãƒ£ãƒ¼æ›¸é¡")

    except Exception as e:
        logging.exception(f"{club_name}: æ›¸é¡â‘¡-2ã®Excelèª­ã¿è¾¼ã¿ã‚¨ãƒ©ãƒ¼")
        issues.append("âŒ æ›¸é¡â‘¡-2ã®èª­ã¿å–ã‚Šã«å¤±æ•—ã—ã¾ã—ãŸã€‚")

    return issues

def check_excel_contents_paper3(club_path, club_name):
    issues = []

    # æ›¸é¡â‘¢ï¼ˆPDFï¼‰ãƒ•ã‚¡ã‚¤ãƒ«æ¢ç´¢
    pattern = os.path.join(club_path, "03_*.pdf")
    files = glob.glob(pattern)

    if not files:
        issues.append("â— æ›¸é¡â‘¢ã€è¦ç´„ãƒ»ä¼šå‰‡ãƒ»å®šæ¬¾ç­‰ã€ãŒè¦‹ã¤ã‹ã‚Šã¾ã›ã‚“ã€‚")
        return issues

    file_path = files[0]  # ç¾æ™‚ç‚¹ã§ã¯æœ€åˆã®ãƒ•ã‚¡ã‚¤ãƒ«ã®ã¿å¯¾è±¡
    try:
        # PDFã®è‡ªå‹•å‡¦ç†ã¯ã¾ã è¡Œã‚ãªã„ã®ã§ã€ãƒ†ãƒ³ãƒ—ãƒ¬ãƒ¼ãƒˆã‚’æ‰‹ä½œæ¥­ãƒã‚§ãƒƒã‚¯ç”¨ã«èª­ã¿è¾¼ã‚€
        from utils.template_loader import load_check_template_paper3
        template = load_check_template_paper3()
        print(f"ğŸ” {club_name}: æ›¸é¡â‘¢ ãƒ†ãƒ³ãƒ—ãƒ¬ãƒ¼ãƒˆä»¶æ•° = {len(template)}")

        # ç¾æ™‚ç‚¹ã§ã¯ãƒ†ãƒ³ãƒ—ãƒ¬ãƒ¼ãƒˆã«æ²¿ã£ãŸè‡ªå‹•ãƒã‚§ãƒƒã‚¯ã¯æœªå¯¾å¿œãªã®ã§ã€ç¢ºèªé …ç›®åã®ã¿å‡ºåŠ›
        for item in template:
            issues.append(f"[æ›¸é¡â‘¢ãƒã‚§ãƒƒã‚¯é …ç›®] {item['ãƒã‚§ãƒƒã‚¯é …ç›®']} â†’ è¦ç¢ºèª")

    except Exception as e:
        logging.exception(f"{club_name}: æ›¸é¡â‘¢ãƒ†ãƒ³ãƒ—ãƒ¬ãƒ¼ãƒˆã®èª­ã¿å–ã‚Šä¸­ã«ã‚¨ãƒ©ãƒ¼")
        issues.append("âŒ æ›¸é¡â‘¢ã®ãƒã‚§ãƒƒã‚¯ãƒ†ãƒ³ãƒ—ãƒ¬ãƒ¼ãƒˆèª­ã¿è¾¼ã¿ã«å¤±æ•—ã—ã¾ã—ãŸã€‚")

    return issues

def check_submissions(clubs, club_master):
    from config.paths import REQUIRED_FILES  # å±€æ‰€çš„ã«ã‚‚ä½¿ãˆã‚‹ã‚ˆã†å†importï¼ˆå¿µã®ãŸã‚ï¼‰
    results = []

    for club_name, club_path in clubs:
        row = {"ã‚¯ãƒ©ãƒ–å": club_name}
        log_messages = []

        try:
            # æ›¸é¡ã®å­˜åœ¨ãƒã‚§ãƒƒã‚¯
            for code, filename in REQUIRED_FILES.items():
                pattern = os.path.join(club_path, f"{code}_*")
                files = glob.glob(pattern)
                if files:
                    row[filename] = "âœ”"
                else:
                    row[filename] = "Ã—"
                    log_messages.append(f"â— {filename} ãŒè¦‹ã¤ã‹ã‚Šã¾ã›ã‚“ã€‚")

            # Excelã®ä¸­èº«ãƒã‚§ãƒƒã‚¯
            if row[REQUIRED_FILES["01"]] == "âœ”":
                log_messages += check_excel_contents_paper1(club_path, club_name, club_master)
            
            if row.get(REQUIRED_FILES["02-1"]) == "âœ”":
                log_messages += check_excel_contents_paper2_1(club_path, club_name)
            
            if row.get(REQUIRED_FILES["02-2"]) == "âœ”":
                log_messages += check_excel_contents_paper2_2(club_path, club_name)
            
            if row.get(REQUIRED_FILES["03"]) == "âœ”":
                log_messages += check_excel_contents_paper3(club_path, club_name)

            # ã‚¯ãƒ©ãƒ–åˆ¥ãƒ­ã‚°å‡ºåŠ›
            save_log(club_name, row, log_messages)
            logging.info(f"{club_name}: ãƒã‚§ãƒƒã‚¯å®Œäº†")

        except Exception as e:
            logging.exception(f"{club_name}: å‡¦ç†ä¸­ã«ã‚¨ãƒ©ãƒ¼ãŒç™ºç”Ÿã—ã¾ã—ãŸ")

        results.append(row)

    return results
